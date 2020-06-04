#!/usr/bin/env python3

'''

Name:           xlscolors.py
Description:    Colorize Excel workbooks using OpenPyxl
Author:         David Paneels

Usage:          see xlscolors.py --help

Stylesheet:

The stylesheet is a YAML file containing foreground and background color information for each keyword to colorize in the workbook.
If a .yaml file with the same name than the Excel file is found we will use it, else we will use the DEFAULT_STYLESHEET file (by default xlscolors.yaml).
Attention: the matching of keywords is NOT case sensitive.

TODO:
    Use another Excel file as reference for stylesheet 
    (analyze each cell, extract unique cell values and according colors, store in dict, and generate a YAML file if needed)


'''

import argparse
import sys
import os
import logging

import yaml
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# Global options
DEFAULT_STYLESHEET = "xlscolors.yaml"


def load_stylesheet(filename=DEFAULT_STYLESHEET):
    '''
    Load stylesheet from YAML file and return headers and keywords dicts
    '''
    if not os.path.exists(filename):
        filename = DEFAULT_STYLESHEET

    # Open YAML configuration file and transform it into a dictionnary
    logging.debug(f"[+] Loading stylesheet from {filename}")
    try:
        with open(filename) as f:
            config_data = yaml.load(f.read(), Loader=yaml.SafeLoader)
    except Exception as e:
        logging.critical(f"[!] Could not load stylesheet {filename} (check YAML syntax)")
        sys.exit(1)

    # Extract the colors definition and keyword-to-color mappings from the config
    headers = config_data.get("headers", {})
    keywords = config_data["keywords"]

    return headers, keywords

def colorize_row_or_cell(ws, cell, fgcolor, bgcolor, bold=False, whole_row=False):
    '''
    Colorize a cell or a whole row depending on the value of whole_row
    '''
    if whole_row:
        colorize_row(ws, cell, fgcolor, bgcolor, bold=False)
    else:
        colorize_cell(cell, fgcolor, bgcolor, bold=False)

def colorize_cell(cell, fgcolor, bgcolor, bold=False):
    '''
    Colorize a single row
    '''
    cell.font = Font(
        color = fgcolor,
        bold = bold
        )
    cell.fill = PatternFill(
        start_color = bgcolor, 
        end_color = bgcolor, 
        fill_type = "solid"
        )                    

def colorize_row(ws, cell, fgcolor, bgcolor, bold):
    '''
    Colorize a whole row based on the index of a cell
    '''
    for cell in ws[cell.row]:
        colorize_cell(cell, fgcolor, bgcolor, bold)

def colorize_worksheet(ws, headers, keywords):
    '''
    Colorize an Excel worksheet 
    '''
    logging.debug(f"[+] Colorizing worksheet {ws}")
    rownum = 0
    for row in ws.iter_rows():
        rownum += 1
        if rownum == 1:
            # Colorize column headers if specified
            if headers:
                colorize_row(
                    ws,
                    cell,
                    headers["fg"], 
                    headers["bg"],
                    bold=headers.get("bold")  
                    )
        else:
            # Colorize the rest
            for cell in row:
                # If the cell is empty just skip to the next one
                if not cell.value:
                    continue
                for keyword, kw_colors in keywords.items():
                    # If the keyword starts/ends with '++' we match anything containing it
                    if keyword.startswith("++") and keyword.endswith("++"):
                        if keyword.strip("++").lower() in str(cell.value).lower():
                            # If whole_row == True -> colorize the entire row if one of the keywords matches
                            # If multiple keywords match, the last one of the list will take precedence
                            colorize_row_or_cell(
                                ws,
                                cell,
                                kw_colors["fg"], 
                                kw_colors["bg"], 
                                bold=kw_colors.get("bold"),
                                whole_row=kw_colors.get("whole_row")
                                )
                    else:
                        # ...else we match the keyword exactly
                        if keyword.lower() == str(cell.value).lower():
                            colorize_row_or_cell(
                                ws,
                                cell,
                                kw_colors["fg"], 
                                kw_colors["bg"], 
                                bold=kw_colors.get("bold"),
                                whole_row=kw_colors.get("whole_row")
                                )

def colorize_workbook(filename, stylesheet="", outfile=""):
    '''
    Colorize a whole Excel workbook
    '''
    # Open workbook
    try:
        wb = load_workbook(filename)
    except:
        logging.critical(f"[!] Could not open {filename}")

    # Check if we a YAML stylesheet with the same name than the Excel file exists, else load the default stylesheet
    if not stylesheet:
        stylesheet = filename.split(".xls")[0] + ".yaml"

    if not os.path.exists(stylesheet):
        stylesheet = DEFAULT_STYLESHEET

    # Load stylesheet from YAML file
    logging.debug(f"[+] Loading stylesheet {stylesheet}")
    headers, keywords = load_stylesheet(stylesheet)

    for ws in wb.worksheets:
        colorize_worksheet(ws, headers, keywords)

    if not outfile:
        outfile = filename

    try:
        wb.save(outfile)
        logging.debug(f"[+] Done writing to {outfile}")
    except:
        logging.critical(f"[!] Could not write to {outfile}")


def main():
    '''
    Main program
    '''
    # Parse command line arguments
    argparser = argparse.ArgumentParser(
        description="Colorize Excel workbooks according to a stylesheet"
        )
    argparser.add_argument(
        "--infile",
        required=True,
        metavar="filename.xlsx",
        help="Excel file to colorize"
        )
    argparser.add_argument(
        "--outfile",
        metavar="filename.xlsx",
        help="save colorized output to file (default=overwrite input file)"
        )
    argparser.add_argument(
        "--stylesheet",
        metavar="filename.yaml",
        help="Stylesheet file in YAML format (default=xlscolors.yaml)",
        )
    argparser.add_argument(
        "--verbose",
        action="store_true",
        help="print additional information to stderr"
        )
    args = argparser.parse_args()

    # Configure logging
    if args.verbose:
        logging.basicConfig(format="%(message)s", level=logging.DEBUG)
    else:
        logging.basicConfig(format="%(message)s", level=logging.INFO)

    # If we don't specify an output file we will overwrite the input file
    if not args.outfile:
        args.outfile = args.infile

    # Colorize the entire workbook according to the stylesheet
    logging.debug(f"[+] Starting colorizing process for {args.infile}...")
    colorize_workbook(args.infile, args.stylesheet, args.outfile)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print()
        sys.exit(1)
    except Exception as e:
        logging.critical(f"[!] An error occured: {str(e)}")
        #raise e
        sys.exit(1)
